import os
import pickle
import re
import statistics
import zipfile
from collections import Counter
import argparse
import importlib.util
import random
import time
import difflib
import glob
import csv
import sys
import unicodedata
import hashlib
import threading
import queue
import traceback
import inspect

import openpyxl

try:
    import sentencepiece as spm
    spm_import_error = None
except Exception as e:
    spm = None
    spm_import_error = str(e)


_HANGUL_RE = re.compile(r"[\uAC00-\uD7A3]")
_HANZI_RE = re.compile(r"[\u4E00-\u9FFF]")
_ASCII_WORD_DIGIT_RE = re.compile(r"[A-Za-z][A-Za-z0-9/_\-\.\+]*|\d+(?:\.\d+)?%?")
_PIECES_RE = re.compile(r"[A-Za-z][A-Za-z0-9/_\-\.\+]*|\d+(?:\.\d+)?%?|[\uAC00-\uD7A3]+|[\u4E00-\u9FFF]+")


def _configure_console_utf8():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    except Exception:
        pass
    try:
        sys.stderr.reconfigure(encoding="utf-8", errors="backslashreplace")
    except Exception:
        pass


def _load_keywords_file(path: str):
    path = os.path.normpath(str(path).strip().strip('"').strip("'"))
    keywords = []
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            keywords.append(line)
    uniq = []
    seen = set()
    for k in keywords:
        k2 = k.lower()
        if k2 in seen:
            continue
        seen.add(k2)
        uniq.append(k)
    
    if not uniq:
        return None
        
    import re
    # Escape keywords and join with OR
    pattern = "|".join(re.escape(k.lower()) for k in uniq)
    try:
        combined_re = re.compile(pattern)
    except Exception:
        combined_re = None
        
    return {"list": uniq, "re": combined_re}


def _has_hangul(text: str) -> bool:
    return bool(text) and bool(re.search(r"[\uAC00-\uD7A3]", text))


def _has_hanzi(text: str) -> bool:
    return bool(text) and bool(re.search(r"[\u4E00-\u9FFF]", text))


def _is_english_only(text: str) -> bool:
    if not isinstance(text, str):
        return False
    s = text.strip()
    if not s:
        return False
    if _has_hangul(s) or _has_hanzi(s):
        return False
    return bool(re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9/_\-\.\+\s]*", s))


def _is_number_only(text: str) -> bool:
    if not isinstance(text, str):
        return False
    s = re.sub(r"\s+", "", text)
    if not s:
        return False
    return bool(re.fullmatch(r"[\d\.,:%\+\-~±]+", s))


def _is_symbol_only(text: str) -> bool:
    if not isinstance(text, str):
        return False
    s = re.sub(r"\s+", "", text)
    if not s:
        return False
    for ch in s:
        if ch.isalnum():
            return False
        cat = unicodedata.category(ch)
        if not (cat and (cat[0] == "P" or cat in ("Sm", "Sc"))):
            return False
    return True


def _pieces_for_single_check(text: str):
    if not isinstance(text, str) or not text:
        return []
    return [p for p in _PIECES_RE.findall(text) if p]


def _ascii_words_digits(text: str):
    if not isinstance(text, str):
        return []
    return _ASCII_WORD_DIGIT_RE.findall(text)


def _hangul_ratio(text: str) -> float:
    if not isinstance(text, str) or not text:
        return 0.0
    h = len(_HANGUL_RE.findall(text))
    return h / max(1, len(text))


def _hanzi_ratio(text: str) -> float:
    if not isinstance(text, str) or not text:
        return 0.0
    h = len(_HANZI_RE.findall(text))
    return h / max(1, len(text))


def _row_hash(ko: str, zh: str) -> bytes:
    s = (ko or "") + "\n" + (zh or "")
    return hashlib.blake2b(s.encode("utf-8", errors="ignore"), digest_size=8).digest()


def score_corpus_xlsx(
    input_xlsx: str,
    output_xlsx: str,
    sheet_name: str | None,
    max_ko_len: int = 128,
    max_zh_len: int = 96,
):
    input_xlsx = os.path.normpath(str(input_xlsx).strip().strip('"').strip("'"))
    output_xlsx = os.path.normpath(str(output_xlsx).strip().strip('"').strip("'"))
    wb_values = None
    wb = None
    try:
        print(f"[score] loading values (read-only): {input_xlsx} ...")
        wb_values = openpyxl.load_workbook(input_xlsx, data_only=True, read_only=True)
        print(f"[score] loading workbook (writeable): {input_xlsx} ...")
        wb = openpyxl.load_workbook(input_xlsx, data_only=False, read_only=False, keep_vba=False)
        print("[score] workbook loaded")
    except zipfile.BadZipFile:
        print(f"\n[Error] 文件损坏或不是有效的 Excel 文件: {input_xlsx}")
        print("请检查文件是否完整，或是否由之前的操作中断导致损坏。")
        raise RuntimeError(f"文件损坏: {input_xlsx}")
    except Exception as e:
        print(f"\n[Error] 无法打开文件 {input_xlsx}: {e}")
        raise

    try:
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        ws_values = wb_values[ws.title] if ws.title in wb_values.sheetnames else wb_values.active
        print("[score] workbooks loaded")

        def get_header(sheet):
            return [c.value for c in sheet[1]]

        header = get_header(ws)
        header_norm = [str(x).strip().lower() if x is not None else "" for x in header]

        def find_col(names, default_1based: int | None = None):
            for n in names:
                n2 = str(n).strip().lower()
                if n2 in header_norm:
                    return header_norm.index(n2) + 1
            return default_1based

        ko_col = find_col(("ko", "korean", "src", "source", "韩文", "韩语"), 2)
        zh_fix_col = find_col(("zh修正", "zh 수정", "zh_fix", "zhfixed", "corrected", "fix", "정답"), 4)
        ko_len_col = find_col(("字符数(ko)", "字符数（ko）", "char_ko", "ko_len"), 5)
        zh_len_col = find_col(("字符数(zh)", "字符数（zh）", "char_zh", "zh_len"), 6)

        score_cols = [
            "len_ratio",
            "over_ko_len",
            "over_zh_len",
            "keep_score",
            "has_hanja",
            "is_duplicate_pair",
            "quality_score",
            "quality_reasons",
        ]

        existing = {str(v).strip(): i + 1 for i, v in enumerate(header) if isinstance(v, str)}
        score_col_idx = {}
        next_col = ws.max_column + 1
        for c in score_cols:
            if c in existing:
                score_col_idx[c] = existing[c]
            else:
                ws.cell(row=1, column=next_col).value = c
                score_col_idx[c] = next_col
                next_col += 1

        def get_text(cell_value):
            return "" if cell_value is None else str(cell_value).strip()

        min_col = min(ko_col, zh_fix_col, ko_len_col, zh_len_col)
        max_col = max(ko_col, zh_fix_col, ko_len_col, zh_len_col)
        off_ko = ko_col - min_col
        off_zh = zh_fix_col - min_col
        off_ko_len = ko_len_col - min_col
        off_zh_len = zh_len_col - min_col

        dup_counter = Counter()
        total_rows = max(0, ws.max_row - 1)
        print("[score] counting duplicates...")
        empty_streak = 0
        rows_iter = ws_values.iter_rows(min_row=2, min_col=min_col, max_col=max_col, values_only=True)
        for i, row in enumerate(rows_iter, start=1):
            if i % 2000 == 0:
                print(f"[score][dup] {i} rows scanned...")
            
            if not row:
                empty_streak += 1
                if empty_streak > 100: break
                continue
                
            ko_s = get_text(row[off_ko] if off_ko < len(row) else None)
            zh_s = get_text(row[off_zh] if off_zh < len(row) else None)
            
            if not ko_s and not zh_s:
                empty_streak += 1
                if empty_streak > 100: break
                continue
            
            empty_streak = 0
            if ko_s and zh_s:
                dup_counter[_row_hash(ko_s, zh_s)] += 1

        sum_score = 0.0
        cnt_score = 0
        over_ko = 0
        over_zh = 0
        bad_ratio = 0
        keep_cnt = 0
        keep_sum = 0.0

        print("[score] scoring rows...")
        empty_streak = 0
        rows_iter = ws_values.iter_rows(min_row=2, min_col=min_col, max_col=max_col, values_only=True)
        for i, row in enumerate(rows_iter, start=1):
            if i % 2000 == 0:
                print(f"[score] {i} rows scored...")
            r = i + 1
            
            if not row:
                empty_streak += 1
                if empty_streak > 100: break
                continue

            ko_s = get_text(row[off_ko] if off_ko < len(row) else None)
            zh_s = get_text(row[off_zh] if off_zh < len(row) else None)
            
            if not ko_s and not zh_s:
                empty_streak += 1
                if empty_streak > 100: break
                continue
                
            empty_streak = 0

            ko_len_raw = row[off_ko_len] if off_ko_len < len(row) else None
            zh_len_raw = row[off_zh_len] if off_zh_len < len(row) else None
            try:
                ko_len = int(ko_len_raw) if ko_len_raw is not None else len(ko_s)
            except Exception:
                ko_len = len(ko_s)
            try:
                zh_len = int(zh_len_raw) if zh_len_raw is not None else len(zh_s)
            except Exception:
                zh_len = len(zh_s)
            lr = (zh_len / max(1, ko_len)) if ko_len else 0.0

            over_ko_len = int(ko_len > int(max_ko_len))
            over_zh_len = int(zh_len > int(max_zh_len))
            over_ko += over_ko_len
            over_zh += over_zh_len

            is_dup = int((ko_s and zh_s) and dup_counter[_row_hash(ko_s, zh_s)] > 1)
            has_hanja = int(_has_hanzi(ko_s))

            keep = None
            src_tokens = set(_ascii_words_digits(ko_s))
            if src_tokens:
                keep_cnt += 1
                out_tokens = set(_ascii_words_digits(zh_s))
                keep = sum(1 for t in src_tokens if t in out_tokens) / max(1, len(src_tokens))
                keep_sum += keep

            score = 100.0
            reasons = []

            if not ko_s or not zh_s:
                score = 0.0
                reasons.append("empty")
            else:
                hr = _hangul_ratio(ko_s)
                zr = _hanzi_ratio(zh_s)
                if hr < 0.2:
                    score -= 20
                    reasons.append("low_hangul_ratio")
                if zr < 0.2:
                    score -= 20
                    reasons.append("low_hanzi_ratio")
                if ko_s == zh_s:
                    score -= 30
                    reasons.append("ko_eq_zh")
                if lr < 0.25 or lr > 3.5:
                    score -= 15
                    reasons.append("len_ratio_out_of_range")
                    bad_ratio += 1
                if keep is not None:
                    score -= 15 * (1.0 - keep)
                    if keep < 1.0:
                        reasons.append("ascii_digit_mismatch")
                if is_dup:
                    score -= 10
                    reasons.append("duplicate_pair")
                if has_hanja:
                    score -= 15
                    reasons.append("ko_has_hanja")

                pieces = _pieces_for_single_check(ko_s)
                if len(pieces) == 1:
                    score -= 8
                    reasons.append("single_like")

                if over_ko_len:
                    score -= 20
                    reasons.append("over_max_ko_len")
                if over_zh_len:
                    score -= 20
                    reasons.append("over_max_zh_len")

            if score < 0:
                score = 0.0
            if score > 100:
                score = 100.0

            sum_score += score
            cnt_score += 1

            ws.cell(row=r, column=score_col_idx["len_ratio"]).value = round(float(lr), 2)
            ws.cell(row=r, column=score_col_idx["over_ko_len"]).value = over_ko_len
            ws.cell(row=r, column=score_col_idx["over_zh_len"]).value = over_zh_len
            ws.cell(row=r, column=score_col_idx["keep_score"]).value = "" if keep is None else round(float(keep), 2)
            ws.cell(row=r, column=score_col_idx["has_hanja"]).value = has_hanja
            ws.cell(row=r, column=score_col_idx["is_duplicate_pair"]).value = is_dup
            ws.cell(row=r, column=score_col_idx["quality_score"]).value = round(float(score), 2)
            ws.cell(row=r, column=score_col_idx["quality_reasons"]).value = "|".join(reasons)

        avg_score = sum_score / max(1, cnt_score)
        keep_avg = (keep_sum / max(1, keep_cnt)) if keep_cnt else 1.0

        try:
            print(f"[score] saving to: {output_xlsx} ...")
            wb.save(output_xlsx)
            print("[score] saved")
        except PermissionError:
            print("\nCannot write file (maybe opened by Excel):", output_xlsx)
            print("Please close Excel and retry.")
            return

        print("\n[score_corpus_xlsx]")
        print("input:", input_xlsx)
        print("sheet:", ws.title)
        print("output:", output_xlsx)
        print("rows:", cnt_score)
        print("avg_quality_score:", round(avg_score, 2))
        print("over_max_ko_len_rate:", round(over_ko / max(1, cnt_score), 6), "max_ko_len:", int(max_ko_len))
        print("over_max_zh_len_rate:", round(over_zh / max(1, cnt_score), 6), "max_zh_len:", int(max_zh_len))
        print("len_ratio_out_of_range_rate:", round(bad_ratio / max(1, cnt_score), 6), "range:", "0.25~3.5")
        print("avg_keep_score(only_when_src_has_ascii_or_digits):", round(keep_avg, 2), "samples:", keep_cnt)
    finally:
        try:
            wb_values.close()
        except Exception:
            pass
        wb.close()


def tag_corpus_xlsx(
    input_xlsx: str,
    output_xlsx: str,
    sheet_name: str | None,
    keywords_path: str,
):
    input_xlsx = os.path.normpath(str(input_xlsx).strip().strip('"').strip("'"))
    output_xlsx = os.path.normpath(str(output_xlsx).strip().strip('"').strip("'"))
    kw_data = _load_keywords_file(keywords_path)
    keywords = kw_data["list"] if kw_data else []
    combined_re = kw_data["re"] if kw_data else None

    tag_cols = [
        "tag",
        "is_single",
        "is_english_only",
        "is_number_only",
        "is_symbol_only",
        "has_hangul",
        "has_hanzi",
        "has_hanja",
        "has_domain_kw",
        "domain_kw_hits",
    ]
    wb = None
    try:
        print(f"[tag] loading workbook: {input_xlsx} ...")
        wb = openpyxl.load_workbook(input_xlsx, data_only=False, read_only=False, keep_vba=False)
        print("[tag] workbook loaded")
    except zipfile.BadZipFile:
        print(f"\n[Error] 文件损坏或不是有效的 Excel 文件: {input_xlsx}")
        print("请检查文件是否完整，或是否由之前的操作中断导致损坏。")
        raise RuntimeError(f"文件损坏: {input_xlsx}")
    except Exception as e:
        print(f"\n[Error] 无法打开文件 {input_xlsx}: {e}")
        raise

    try:
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        header = [c.value for c in ws[1]]
        header_norm = [str(x).strip().lower() if x is not None else "" for x in header]

        def find_col(names, default_1based: int | None = None):
            for n in names:
                n2 = str(n).strip().lower()
                if n2 in header_norm:
                    return header_norm.index(n2) + 1
            return default_1based

        ko_col = find_col(("ko", "korean", "src", "source", "韩文", "韩语"), 2)
        zh_fix_col = find_col(("zh修正", "zh 수정", "zh_fix", "zhfixed", "corrected", "fix", "정답"), 4)

        existing = {str(v).strip(): i + 1 for i, v in enumerate(header) if isinstance(v, str)}
        tag_col_idx = {}
        next_col = ws.max_column + 1
        for c in tag_cols:
            if c in existing:
                tag_col_idx[c] = existing[c]
            else:
                ws.cell(row=1, column=next_col).value = c
                tag_col_idx[c] = next_col
                next_col += 1

        def get_text(v):
            return "" if v is None else str(v).strip()

        tag_counter = Counter()
        total_rows = max(0, ws.max_row - 1)
        print(f"[tag] max_row reported by excel: {ws.max_row}")
        
        # Determine max_col for iter_rows
        max_col_to_read = max(ko_col, zh_fix_col)
        
        # Use iter_rows to avoid cell-by-cell overhead and stop on empty rows
        rows_iter = ws.iter_rows(min_row=2, min_col=1, max_col=max_col_to_read, values_only=True)
        empty_streak = 0
        for i, row in enumerate(rows_iter, start=1):
            r = i + 1
            if i % 2000 == 0:
                print(f"[tag] {i} rows processed...")
            
            if not row:
                empty_streak += 1
                if empty_streak > 100: break # Stop if 100 consecutive rows are truly empty
                continue
            
            ko_s = get_text(row[ko_col-1] if ko_col-1 < len(row) else None)
            zh_s = get_text(row[zh_fix_col-1] if zh_fix_col-1 < len(row) else None)

            if not ko_s and not zh_s:
                empty_streak += 1
                if empty_streak > 100: break
                continue
            
            empty_streak = 0 # Reset streak if we find data

            has_h = _has_hangul(ko_s)
            has_z = _has_hanzi(zh_s)
            has_hanja = _has_hanzi(ko_s)
            is_en = _is_english_only(ko_s)
            is_num = _is_number_only(ko_s)
            is_sym = _is_symbol_only(ko_s)
            pieces = _pieces_for_single_check(ko_s)
            is_single = int(len(pieces) == 1)

            hits = []
            ko_low = ko_s.lower()
            if combined_re and combined_re.search(ko_low):
                for k in keywords:
                    if k.lower() in ko_low:
                        hits.append(k)
                        if len(hits) >= 8:
                            break
            has_kw = int(len(hits) > 0)

            if not ko_s or not zh_s:
                tag = "empty"
            elif is_en and not has_h:
                tag = "english"
            elif is_sym:
                tag = "symbol"
            elif is_num:
                tag = "number"
            elif is_single:
                tag = "single_word"
            elif has_kw:
                tag = "automation"
            elif has_h:
                tag = "life"
            else:
                tag = "other"

            tag_counter[tag] += 1
            ws.cell(row=r, column=tag_col_idx["tag"]).value = tag
            ws.cell(row=r, column=tag_col_idx["is_single"]).value = is_single
            ws.cell(row=r, column=tag_col_idx["is_english_only"]).value = int(is_en)
            ws.cell(row=r, column=tag_col_idx["is_number_only"]).value = int(is_num)
            ws.cell(row=r, column=tag_col_idx["is_symbol_only"]).value = int(is_sym)
            ws.cell(row=r, column=tag_col_idx["has_hangul"]).value = int(has_h)
            ws.cell(row=r, column=tag_col_idx["has_hanzi"]).value = int(has_z)
            ws.cell(row=r, column=tag_col_idx["has_hanja"]).value = int(has_hanja)
            ws.cell(row=r, column=tag_col_idx["has_domain_kw"]).value = has_kw
            ws.cell(row=r, column=tag_col_idx["domain_kw_hits"]).value = "|".join(hits)

        try:
            print(f"[tag] saving to: {output_xlsx} ...")
            wb.save(output_xlsx)
            print("[tag] saved")
        except PermissionError:
            print("\nCannot write file (maybe opened by Excel):", output_xlsx)
            print("Please close Excel and retry.")
            return

        print("\n[tag_corpus_xlsx]")
        print("input:", input_xlsx)
        print("sheet:", ws.title)
        print("output:", output_xlsx)
        print("keywords_file:", os.path.normpath(keywords_path))
        print("\n[tag counts]")
        for k, v in tag_counter.most_common():
            print(f"{k}: {v}")
    finally:
        wb.close()


def summarize_corpus_xlsx(input_xlsx: str, sheet_name: str | None = None) -> None:
    input_xlsx = os.path.normpath(str(input_xlsx).strip().strip('"').strip("'"))
    if not os.path.exists(input_xlsx):
        raise RuntimeError(f"找不到文件: {input_xlsx}")

    wb = None
    try:
        print(f"[summary] loading workbook: {input_xlsx} ...")
        wb = openpyxl.load_workbook(input_xlsx, data_only=True, read_only=True)
        print("[summary] workbook loaded")
    except zipfile.BadZipFile:
        print(f"\n[Error] 文件损坏或不是有效的 Excel 文件: {input_xlsx}")
        print("请检查文件是否完整，或是否由之前的操作中断导致损坏。")
        raise RuntimeError(f"文件损坏: {input_xlsx}")
    except Exception as e:
        print(f"\n[Error] 无法打开文件 {input_xlsx}: {e}")
        raise

    try:
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        header = [c.value for c in ws[1]]
        header_norm = [str(x).strip().lower() if x is not None else "" for x in header]

        def find_col(names, default_1based: int | None = None):
            for n in names:
                n2 = str(n).strip().lower()
                if n2 in header_norm:
                    return header_norm.index(n2) + 1
            return default_1based

        ko_col = find_col(("ko", "korean", "src", "source", "韩文", "韩语"), None)
        zh_fix_col = find_col(("zh修正", "zh 수정", "zh_fix", "zhfixed", "corrected", "fix", "정답", "zh", "中文"), None)
        tag_col = find_col(("tag",), None)
        has_kw_col = find_col(("has_domain_kw",), None)
        has_hanja_col = find_col(("has_hanja",), None)

        score_cols = {
            "len_ratio": find_col(("len_ratio",), None),
            "over_ko_len": find_col(("over_ko_len",), None),
            "over_zh_len": find_col(("over_zh_len",), None),
            "keep_score": find_col(("keep_score",), None),
            "is_duplicate_pair": find_col(("is_duplicate_pair",), None),
            "quality_score": find_col(("quality_score",), None),
        }

        def get_text(v):
            return "" if v is None else str(v).strip()

        def to_float(v):
            try:
                if v is None or v == "":
                    return None
                return float(v)
            except Exception:
                return None

        def to_int01(v):
            try:
                return int(v) != 0
            except Exception:
                return False

        tag_counter = Counter()
        total_ko = 0
        total_zh = 0
        qs = []
        keep = []
        lr = []
        over_ko = 0
        over_zh = 0
        dup = 0
        has_kw = 0
        has_hanja_cnt = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            ko = get_text(row[ko_col - 1]) if ko_col and ko_col - 1 < len(row) else ""
            if not ko:
                continue
            total_ko += 1

            zh = get_text(row[zh_fix_col - 1]) if zh_fix_col and zh_fix_col - 1 < len(row) else ""
            if zh:
                total_zh += 1

            if tag_col and tag_col - 1 < len(row):
                tag = get_text(row[tag_col - 1])
                if tag:
                    tag_counter[tag] += 1

            c = score_cols.get("quality_score")
            if c and c - 1 < len(row):
                v = to_float(row[c - 1])
                if v is not None:
                    qs.append(v)

            c = score_cols.get("keep_score")
            if c and c - 1 < len(row):
                v = to_float(row[c - 1])
                if v is not None:
                    keep.append(v)

            c = score_cols.get("len_ratio")
            if c and c - 1 < len(row):
                v = to_float(row[c - 1])
                if v is not None:
                    lr.append(v)

            c = score_cols.get("over_ko_len")
            if c and c - 1 < len(row):
                over_ko += int(to_int01(row[c - 1]))

            c = score_cols.get("over_zh_len")
            if c and c - 1 < len(row):
                over_zh += int(to_int01(row[c - 1]))

            c = score_cols.get("is_duplicate_pair")
            if c and c - 1 < len(row):
                dup += int(to_int01(row[c - 1]))

            if has_kw_col and has_kw_col - 1 < len(row):
                has_kw += int(to_int01(row[has_kw_col - 1]))

            if has_hanja_col and has_hanja_col - 1 < len(row):
                has_hanja_cnt += int(to_int01(row[has_hanja_col - 1]))

        def pct(x, den):
            return round((x / max(1, den)) * 100.0, 2)

        def qstat(a: list[float]):
            if not a:
                return None
            a = sorted(a)
            n = len(a)

            def q(p):
                k = (n - 1) * p
                f = int(math.floor(k))
                c = int(math.ceil(k))
                if f == c:
                    return a[f]
                return a[f] * (c - k) + a[c] * (k - f)

            return {
                "avg": round(statistics.mean(a), 2),
                "p50": round(q(0.5), 2),
                "p10": round(q(0.1), 2),
                "p90": round(q(0.9), 2),
                "min": round(a[0], 2),
                "max": round(a[-1], 2),
            }

        print("\n[Corpus Summary]")
        print("file:", input_xlsx)
        print("sheet:", ws.title)
        print("rows_with_ko:", total_ko)
        print("zh_fill_rate_%:", pct(total_zh, total_ko))
        print("tag_counts_top10:", tag_counter.most_common(10))
        print("quality_score:", qstat(qs))
        print("quality_ge_80_%:", pct(sum(1 for v in qs if v >= 80), len(qs)))
        print("quality_ge_90_%:", pct(sum(1 for v in qs if v >= 90), len(qs)))
        print("keep_score:", qstat(keep))
        print("len_ratio:", qstat(lr))
        print("len_ratio_out_of_range_%:", pct(sum(1 for v in lr if v < 0.25 or v > 3.5), len(lr)))
        print("over_ko_len_%:", pct(over_ko, total_ko))
        print("over_zh_len_%:", pct(over_zh, total_ko))
        print("duplicate_pair_%:", pct(dup, total_ko))
        print("has_domain_kw_%:", pct(has_kw, total_ko))
        print("has_hanja_%:", pct(has_hanja_cnt, total_ko))
    finally:
        wb.close()


def clean_text(sentence: str) -> str:
    if not isinstance(sentence, str):
        return ""
    sentence = re.sub(r"[^\w\s\uAC00-\uD7A3\u4e00-\u9fa5]", "", sentence)
    return sentence.strip()


def try_build_tokenizers():
    ko_name = "split"
    zh_name = "char"

    def tok_ko(s: str):
        return s.split()

    def tok_zh(s: str):
        return list(s)

    try:
        from konlpy.tag import Okt

        okt = Okt()

        def tok_ko(s: str):
            return okt.morphs(s)

        ko_name = "Okt"
    except Exception as e:
        ko_name = f"split (Okt unavailable: {e})"

    try:
        import jieba

        def tok_zh(s: str):
            return jieba.lcut(s)

        zh_name = "jieba"
    except Exception as e:
        zh_name = f"char (jieba unavailable: {e})"

    return tok_ko, tok_zh, ko_name, zh_name


def summarize_lengths(name: str, seqs):
    lens = [len(x) for x in seqs]
    lens_sorted = sorted(lens)

    def pct(p: int):
        idx = int(round((p / 100) * (len(lens_sorted) - 1)))
        return lens_sorted[idx]

    print(f"\n[{name}] length (tokens)")
    print("min/mean/median/max:", min(lens), sum(lens) / len(lens), statistics.median(lens), max(lens))
    for p in (50, 75, 90, 95, 99):
        print(f"p{p}:", pct(p))
    for th in (50, 80, 100, 150, 200):
        over = sum(1 for L in lens if L > th)
        print(f"> {th}: {over} ({over/len(lens)*100:.2f}%)")


def unk_report(name: str, tok_seqs, vocab, topn: int = 30):
    total = 0
    unk_count = 0
    sent_with_unk = 0
    unk_counter = Counter()

    for toks in tok_seqs:
        has_unk = False
        for t in toks:
            total += 1
            if t not in vocab:
                unk_count += 1
                unk_counter[t] += 1
                has_unk = True
        if has_unk:
            sent_with_unk += 1

    print(f"\n[{name}] vocab coverage")
    print("total tokens:", total)
    print("unk tokens:", unk_count, f"({(unk_count/total*100) if total else 0:.2f}%)")
    print("sentences with any unk:", sent_with_unk, f"({sent_with_unk/len(tok_seqs)*100:.2f}%)")
    print("top unk tokens:")
    for t, c in unk_counter.most_common(topn):
        print(f"  {t!r}: {c}")


def _read_parallel_corpus_xlsx(xlsx_path: str, max_rows: int | None = None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    ko_sents, zh_sents = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or len(row) < 4:
            continue
        ko, zh = row[1], row[3]
        if ko and zh:
            ko_sents.append(str(ko))
            zh_sents.append(str(zh))
        if max_rows and len(ko_sents) >= max_rows:
            break
    wb.close()
    return ko_sents, zh_sents


def vocab_diagnose(root: str, corpus_xlsx: str, model_dir: str):
    corpus_path = os.path.join(root, corpus_xlsx)
    model_path = None
    try:
        model_path, ko_vocab_path, zh_vocab_path = _infer_model_paths(model_dir)
    except Exception as e:
        print("Failed to locate model/vocab under model_dir:", model_dir)
        print("Tip: pass --model-dir to a specific epoch folder that contains .pth/.pkl files.")
        raise

    print("corpus_path:", corpus_path, "exists=", os.path.exists(corpus_path))
    print("model_path:", model_path, "exists=", os.path.exists(model_path))
    print("ko_vocab_path:", ko_vocab_path, "exists=", os.path.exists(ko_vocab_path))
    print("zh_vocab_path:", zh_vocab_path, "exists=", os.path.exists(zh_vocab_path))

    with open(ko_vocab_path, "rb") as f:
        ko_vocab = pickle.load(f)
    with open(zh_vocab_path, "rb") as f:
        zh_vocab = pickle.load(f)

    print("ko_vocab size:", len(ko_vocab))
    print("zh_vocab size:", len(zh_vocab))

    tok_ko, tok_zh, ko_name, zh_name = try_build_tokenizers()
    print("tokenizer ko:", ko_name)
    print("tokenizer zh:", zh_name)

    ko_sents, zh_sents = _read_parallel_corpus_xlsx(corpus_path, max_rows=None)
    print("pairs read:", len(ko_sents))
    ko_sents = [clean_text(s) for s in ko_sents]
    zh_sents = [clean_text(s) for s in zh_sents]
    nonempty = [(k, z) for k, z in zip(ko_sents, zh_sents) if k and z]
    ko_sents = [k for k, _ in nonempty]
    zh_sents = [z for _, z in nonempty]
    print("pairs after clean+drop empty:", len(ko_sents))

    print("tokenizing...")
    ko_tok = [tok_ko(s) for s in ko_sents]
    zh_tok = [tok_zh(s) for s in zh_sents]

    summarize_lengths("KO", ko_tok)
    summarize_lengths("ZH", zh_tok)

    unk_report("KO", ko_tok, ko_vocab)
    unk_report("ZH", zh_tok, zh_vocab)


def _load_module_from_path(module_name: str, file_path: str):
    spec = importlib.util.spec_from_file_location(module_name, file_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Failed to load module: {module_name} from {file_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module

def _infer_model_paths(model_dir: str, prefer_tag: str | None = None):
    model_dir = os.path.normpath(str(model_dir).strip().strip('"').strip("'"))
    if not os.path.isdir(model_dir):
        raise RuntimeError(f"model_dir 不是目录: {model_dir}")

    preferred = []
    if prefer_tag:
        preferred.append(prefer_tag)
    preferred += ["v3_2_1", "v3_0_1", "v3_2", "v3_attn", "best_model"]

    candidates = glob.glob(os.path.join(model_dir, "**", "*.pth"), recursive=True)
    best_model = None
    best_score = (-1, -1.0)
    for p in candidates:
        base = os.path.basename(p).lower()
        if "best_model" not in base and not base.endswith(".pth"):
            continue
        score0 = 0
        for i, tag in enumerate(preferred[::-1]):
            if tag and tag.lower() in base:
                score0 = max(score0, i + 1)
        score = (score0, os.path.getmtime(p))
        if score > best_score:
            best_score = score
            best_model = p

    if not best_model:
        raise RuntimeError(f"在目录中找不到 .pth 模型: {model_dir}")

    ko_vocab = None
    zh_vocab = None
    ko_cands = glob.glob(os.path.join(os.path.dirname(best_model), "**", "*ko*vocab*.pkl"), recursive=True) + glob.glob(
        os.path.join(model_dir, "**", "*ko*vocab*.pkl"), recursive=True
    )
    zh_cands = glob.glob(os.path.join(os.path.dirname(best_model), "**", "*zh*vocab*.pkl"), recursive=True) + glob.glob(
        os.path.join(model_dir, "**", "*zh*vocab*.pkl"), recursive=True
    )

    def pick_vocab(cands, prefer):
        best = None
        best_sc = (-1, -1.0)
        for p in cands:
            base = os.path.basename(p).lower()
            sc0 = 0
            for i, tag in enumerate(prefer[::-1]):
                if tag and tag.lower() in base:
                    sc0 = max(sc0, i + 1)
            sc = (sc0, os.path.getmtime(p))
            if sc > best_sc:
                best_sc = sc
                best = p
        return best

    ko_vocab = pick_vocab(ko_cands, preferred)
    zh_vocab = pick_vocab(zh_cands, preferred)
    if not ko_vocab or not zh_vocab:
        raise RuntimeError(f"找到模型但没找到 vocab：model={best_model}, ko_vocab={ko_vocab}, zh_vocab={zh_vocab}")

    return best_model, ko_vocab, zh_vocab


def _read_eval_set(eval_path: str, sheet: str | None = None):
    eval_path = os.path.normpath(str(eval_path).strip().strip('"').strip("'"))
    if not os.path.exists(eval_path):
        raise RuntimeError(f"找不到 eval-set 文件: {eval_path}")

    items = []
    ext = os.path.splitext(eval_path)[1].lower()
    if ext in (".txt", ".tsv"):
        with open(eval_path, "r", encoding="utf-8") as f:
            for raw in f:
                line = raw.strip()
                if not line or line.startswith("#"):
                    continue
                if "\t" in line:
                    ko, ref = line.split("\t", 1)
                    items.append({"ko": ko.strip(), "ref": ref.strip()})
                else:
                    items.append({"ko": line, "ref": ""})
        return items

    if ext == ".csv":
        with open(eval_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            rows = list(reader)
        if not rows:
            return items
        header = [str(x).strip().lower() for x in rows[0]]
        ko_idx = None
        ref_idx = None
        for i, h in enumerate(header):
            if h in ("ko", "korean", "source", "src", "韩文", "韩语"):
                ko_idx = i
            if h in ("zh", "chinese", "target", "tgt", "ref", "中文"):
                ref_idx = i
        start_row = 1 if ko_idx is not None else 0
        for r in rows[start_row:]:
            if not r:
                continue
            ko = r[ko_idx] if ko_idx is not None and ko_idx < len(r) else r[0]
            ref = r[ref_idx] if ref_idx is not None and ref_idx < len(r) else ""
            ko = "" if ko is None else str(ko).strip()
            if not ko:
                continue
            items.append({"ko": ko, "ref": "" if ref is None else str(ref).strip()})
        return items

    if ext in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(eval_path, data_only=True, read_only=True)
        try:
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            header = [str(x).strip().lower() if x is not None else "" for x in (header_row or [])]
            ko_idx = None
            ref_idx = None
            for i, h in enumerate(header):
                if h in ("ko", "korean", "source", "src", "韩文", "韩语"):
                    ko_idx = i
                if h in ("zh", "chinese", "target", "tgt", "ref", "中文"):
                    ref_idx = i
            has_header = ko_idx is not None
            if not has_header:
                if header_row:
                    ko = header_row[0] if len(header_row) > 0 else ""
                    ref = header_row[1] if len(header_row) > 1 else ""
                    ko = "" if ko is None else str(ko).strip()
                    if ko:
                        items.append({"ko": ko, "ref": "" if ref is None else str(ref).strip()})
            for row in rows_iter:
                if not row:
                    continue
                ko = row[ko_idx] if ko_idx is not None and ko_idx < len(row) else row[0]
                ref = row[ref_idx] if ref_idx is not None and ref_idx < len(row) else ""
                ko = "" if ko is None else str(ko).strip()
                if not ko:
                    continue
                items.append({"ko": ko, "ref": "" if ref is None else str(ref).strip()})
            return items
        finally:
            wb.close()

    raise RuntimeError(f"不支持的 eval-set 格式: {ext}，请用 .txt/.tsv/.csv/.xlsx")


def _repetition_bigram_ratio(text: str):
    if not isinstance(text, str):
        return 0.0, 0
    s = re.sub(r"\s+", "", text)
    if len(s) < 2:
        return 0.0, len(s)
    bigrams = [s[i : i + 2] for i in range(len(s) - 1)]
    rep_ratio = 1.0 - (len(set(bigrams)) / max(1, len(bigrams)))
    max_run = 1
    cur = 1
    for i in range(1, len(s)):
        if s[i] == s[i - 1]:
            cur += 1
            if cur > max_run:
                max_run = cur
        else:
            cur = 1
    return rep_ratio, max_run


def eval_set_word(
    root: str,
    eval_path: str,
    model_dir: str,
    translator_path: str,
    eval_out: str | None = None,
    eval_sheet: str | None = None,
    n: int | None = None,
    seed: int = 42,
    max_len: int = 50,
):
    try:
        import torch
    except Exception as e:
        print("缺少 torch，无法在本环境跑评估。错误:", e)
        print("请使用装有 torch 的解释器运行，比如：python diagnose_mt.py --eval-set xxx.txt")
        return

    eval_items = _read_eval_set(eval_path, sheet=eval_sheet)
    if not eval_items:
        print("eval-set 为空，未评估。")
        return

    if n is not None and n > 0 and len(eval_items) > n:
        random.seed(seed)
        random.shuffle(eval_items)
        eval_items = eval_items[:n]

    mod = _load_module_from_path("rt_eval", translator_path)
    user_dict_path = os.path.join(root, "user_dict.md")
    user_dict = mod.load_user_dict(user_dict_path) if hasattr(mod, "load_user_dict") else ({}, {}, {}, {}, set())

    model_path, ko_vocab_path, zh_vocab_path = _infer_model_paths(model_dir)
    with open(ko_vocab_path, "rb") as f:
        ko_vocab = pickle.load(f)
    with open(zh_vocab_path, "rb") as f:
        zh_vocab = pickle.load(f)

    device = torch.device("cpu")
    input_dim = len(ko_vocab)
    output_dim = len(zh_vocab)
    enc_emb_dim = 256
    dec_emb_dim = 256
    hid_dim = 512
    n_layers = 1

    if not all(hasattr(mod, x) for x in ("Attention", "Encoder", "Decoder", "Seq2Seq")):
        raise RuntimeError(f"translator 脚本缺少模型类: {translator_path}")

    attn = mod.Attention(hid_dim)
    enc = mod.Encoder(input_dim, enc_emb_dim, hid_dim, n_layers, 0)
    dec = mod.Decoder(output_dim, dec_emb_dim, hid_dim, n_layers, 0, attn)
    model = mod.Seq2Seq(enc, dec, device).to(device)
    model.load_state_dict(torch.load(model_path, map_location=device))
    model.eval()

    t0 = time.perf_counter()
    rows = []
    for it in eval_items:
        ko = it.get("ko", "")
        ref = it.get("ref", "")
        try:
            out = mod.translate_sentence(ko, model, ko_vocab, zh_vocab, device, user_dict, max_len=max_len, show_tokens=False)
        except Exception as e:
            print(f"[eval] 翻译异常: {type(e).__name__}: {e}", flush=True)
            print("KO:", ko, flush=True)
            out = "[ERROR]"

        keep_src = _ascii_words_digits(ko)
        keep_out = _ascii_words_digits(out)
        keep_src_set = set(keep_src)
        keep_score = sum(1 for x in keep_src_set if x in set(keep_out)) / max(1, len(keep_src_set))

        rep_ratio, max_run = _repetition_bigram_ratio(out)
        unk = "<unk>" in (out or "")
        out_has_q = "?" in (out or "")
        src_has_q = "?" in (ko or "")

        src_len = len(re.sub(r"\s+", "", str(ko)))
        out_len = len(re.sub(r"\s+", "", str(out)))
        len_ratio_src = out_len / max(1, src_len)
        ref_len = len(re.sub(r"\s+", "", str(ref))) if ref else 0
        len_ratio_ref = (out_len / max(1, ref_len)) if ref_len else None

        rows.append(
            {
                "ko": ko,
                "ref": ref,
                "out": out,
                "unk": int(unk),
                "src_qmark": int(src_has_q),
                "out_qmark": int(out_has_q),
                "keep_score": keep_score,
                "rep_bigram": rep_ratio,
                "max_run": max_run,
                "src_len": src_len,
                "out_len": out_len,
                "len_ratio_src": len_ratio_src,
                "len_ratio_ref": len_ratio_ref,
            }
        )
    t1 = time.perf_counter()

    print("\n[Eval 总结]")
    print("translator:", translator_path)
    print("model:", model_path)
    print("ko_vocab:", ko_vocab_path)
    print("zh_vocab:", zh_vocab_path)
    print("samples:", len(rows))
    print("elapsed_sec:", round(t1 - t0, 3))
    print("unk rate:", sum(r["unk"] for r in rows) / max(1, len(rows)))
    print("out '?' rate:", sum(r["out_qmark"] for r in rows) / max(1, len(rows)))
    print("avg keep(ascii/digit) score:", sum(r["keep_score"] for r in rows) / max(1, len(rows)))
    print("avg rep_bigram:", sum(r["rep_bigram"] for r in rows) / max(1, len(rows)))
    print("avg len_ratio(src):", sum(r["len_ratio_src"] for r in rows) / max(1, len(rows)))

    def collapse_flag(r):
        return r["out_len"] <= 1 or r["len_ratio_src"] < 0.25

    print("collapse rate:", sum(1 for r in rows if collapse_flag(r)) / max(1, len(rows)))

    rows_sorted = sorted(rows, key=lambda r: (collapse_flag(r), r["unk"], r["rep_bigram"], -r["keep_score"]), reverse=True)
    show = min(20, len(rows_sorted))
    print(f"\n[问题样本 Top {show}]")
    for i in range(show):
        r = rows_sorted[i]
        print("\n---")
        print("KO:", r["ko"])
        if r["ref"]:
            print("REF:", r["ref"])
        print("OUT:", r["out"])
        print(
            "unk:",
            r["unk"],
            "keep:",
            round(r["keep_score"], 2),
            "rep_bigram:",
            round(r["rep_bigram"], 2),
            "max_run:",
            r["max_run"],
            "len_ratio_src:",
            round(r["len_ratio_src"], 2),
        )

    if eval_out:
        eval_out = os.path.normpath(str(eval_out).strip().strip('"').strip("'"))
        out_wb = openpyxl.Workbook()
        ws = out_wb.active
        ws.title = "Eval"
        ws.append(
            [
                "ko",
                "ref",
                "out",
                "unk",
                "src_qmark",
                "out_qmark",
                "keep_score",
                "rep_bigram",
                "max_run",
                "src_len",
                "out_len",
                "len_ratio_src",
                "len_ratio_ref",
            ]
        )
        for r in rows:
            ws.append(
                [
                    r["ko"],
                    r["ref"],
                    r["out"],
                    r["unk"],
                    r["src_qmark"],
                    r["out_qmark"],
                    float(r["keep_score"]),
                    float(r["rep_bigram"]),
                    int(r["max_run"]),
                    int(r["src_len"]),
                    int(r["out_len"]),
                    float(r["len_ratio_src"]),
                    "" if r["len_ratio_ref"] is None else float(r["len_ratio_ref"]),
                ]
            )
        out_wb.save(eval_out)
        print("\n已保存:", eval_out)


def _unwrap_state_dict_plain(state):
    if not isinstance(state, dict):
        return state
    if any(isinstance(k, str) and k.startswith("module.") for k in state.keys()):
        return {k[7:]: v for k, v in state.items() if isinstance(k, str)}
    return state


def _infer_n_layers_from_state_dict_plain(state: dict) -> int:
    if not isinstance(state, dict):
        return 2
    max_idx = -1
    for k in state.keys():
        if not isinstance(k, str):
            continue
        m = re.match(r"^encoder\.rnn\.weight_ih_l(\d+)$", k)
        if m:
            max_idx = max(max_idx, int(m.group(1)))
    if max_idx >= 0:
        return max_idx + 1
    for k in state.keys():
        if not isinstance(k, str):
            continue
        m = re.match(r"^decoder\.rnn\.weight_ih_l(\d+)$", k)
        if m:
            max_idx = max(max_idx, int(m.group(1)))
    return (max_idx + 1) if max_idx >= 0 else 2


def _resolve_bpe_model_paths(model_dir: str) -> tuple[str, str, str, str]:
    model_dir = os.path.normpath(str(model_dir).strip().strip('"').strip("'"))
    if os.path.isfile(model_dir) and model_dir.lower().endswith(".pth"):
        model_path = model_dir
        model_dir = os.path.dirname(model_path)
    else:
        model_path = os.path.join(model_dir, "best_model_v3_2_1_bpe_attn.pth")
    ko_spm_path = os.path.join(model_dir, "spm_ko_v3_2_1.model")
    zh_spm_path = os.path.join(model_dir, "spm_zh_v3_2_1.model")

    if (not os.path.exists(model_path)) and os.path.isdir(model_dir):
        cands = glob.glob(os.path.join(model_dir, "**", "best_model_v3_2_1_bpe_attn.pth"), recursive=True)
        if cands:
            model_path = sorted(cands, key=lambda p: os.path.getmtime(p))[-1]
            model_dir = os.path.dirname(model_path)
            ko_spm_path = os.path.join(model_dir, "spm_ko_v3_2_1.model")
            zh_spm_path = os.path.join(model_dir, "spm_zh_v3_2_1.model")

    if (not os.path.exists(ko_spm_path)) and os.path.isdir(model_dir):
        cands = glob.glob(os.path.join(model_dir, "**", "spm_ko_v3_2_1.model"), recursive=True)
        if cands:
            ko_spm_path = sorted(cands, key=lambda p: os.path.getmtime(p))[-1]
    if (not os.path.exists(zh_spm_path)) and os.path.isdir(model_dir):
        cands = glob.glob(os.path.join(model_dir, "**", "spm_zh_v3_2_1.model"), recursive=True)
        if cands:
            zh_spm_path = sorted(cands, key=lambda p: os.path.getmtime(p))[-1]

    return model_dir, model_path, ko_spm_path, zh_spm_path


def eval_set_bpe(
    root: str,
    eval_path: str,
    model_dir: str,
    translator_path: str,
    eval_out: str | None = None,
    eval_sheet: str | None = None,
    n: int | None = None,
    seed: int = 42,
    max_len: int = 50,
):
    try:
        import torch
    except Exception as e:
        print("缺少 torch，无法在本环境跑评估。错误:", e)
        print("请使用装有 torch 的解释器运行，比如：python diagnose_mt.py --eval-set xxx.txt")
        return

    if spm is None:
        print("缺少 sentencepiece，无法在本环境跑 BPE 评估。错误:", spm_import_error)
        print("请使用装有 sentencepiece 的解释器运行。")
        return

    eval_items = _read_eval_set(eval_path, sheet=eval_sheet)
    if not eval_items:
        print("eval-set 为空，未评估。")
        return

    if n is not None and n > 0 and len(eval_items) > n:
        random.seed(seed)
        random.shuffle(eval_items)
        eval_items = eval_items[:n]

    mod = _load_module_from_path("rt_eval_bpe", translator_path)
    user_dict_path = os.path.join(root, "user_dict.md")
    user_dict = mod.load_user_dict(user_dict_path) if hasattr(mod, "load_user_dict") else ({}, {}, {}, {})

    model_dir2, model_path, ko_spm_path, zh_spm_path = _resolve_bpe_model_paths(model_dir)
    if not os.path.exists(model_path):
        raise RuntimeError(f"找不到 BPE 模型文件: {model_path}")
    if not os.path.exists(ko_spm_path) or not os.path.exists(zh_spm_path):
        raise RuntimeError(f"找不到 SentencePiece 模型: {ko_spm_path} / {zh_spm_path}")

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    ko_sp = spm.SentencePieceProcessor(model_file=ko_spm_path)
    zh_sp = spm.SentencePieceProcessor(model_file=zh_spm_path)

    input_dim = int(ko_sp.get_piece_size())
    output_dim = int(zh_sp.get_piece_size())
    enc_emb_dim = 256
    dec_emb_dim = 256
    hid_dim = 512

    if not all(hasattr(mod, x) for x in ("Attention", "Encoder", "Decoder", "Seq2Seq")):
        raise RuntimeError(f"translator 脚本缺少模型类: {translator_path}")

    state_obj = torch.load(model_path, map_location=device)
    if isinstance(state_obj, dict) and "state_dict" in state_obj and isinstance(state_obj["state_dict"], dict):
        state = state_obj["state_dict"]
    elif isinstance(state_obj, dict) and "model_state_dict" in state_obj and isinstance(state_obj["model_state_dict"], dict):
        state = state_obj["model_state_dict"]
    else:
        state = state_obj
    state = _unwrap_state_dict_plain(state)

    infer_fn = getattr(mod, "infer_n_layers_from_state_dict", None) or getattr(mod, "infer_n_layers_from_state_dict_plain", None)
    if callable(infer_fn):
        try:
            n_layers = int(infer_fn(state))
        except Exception:
            n_layers = _infer_n_layers_from_state_dict_plain(state if isinstance(state, dict) else {})
    else:
        n_layers = _infer_n_layers_from_state_dict_plain(state if isinstance(state, dict) else {})

    attn = mod.Attention(hid_dim)
    enc = mod.Encoder(input_dim, enc_emb_dim, hid_dim, n_layers, 0).to(device)
    dec = mod.Decoder(output_dim, dec_emb_dim, hid_dim, n_layers, 0, attn).to(device)
    model = mod.Seq2Seq(enc, dec, device).to(device)
    model.load_state_dict(state)
    model.eval()

    t0 = time.perf_counter()
    rows = []
    for it in eval_items:
        ko = it.get("ko", "")
        ref = it.get("ref", "")
        try:
            out = mod.translate_sentence(ko, model, ko_sp, zh_sp, device, user_dict, max_len=max_len)
        except Exception as e:
            print(f"[eval] 翻译异常: {type(e).__name__}: {e}", flush=True)
            print("KO:", ko, flush=True)
            out = "[ERROR]"

        keep_src = _ascii_words_digits(ko)
        keep_out = _ascii_words_digits(out)
        keep_src_set = set(keep_src)
        keep_score = sum(1 for x in keep_src_set if x in set(keep_out)) / max(1, len(keep_src_set))

        rep_ratio, max_run = _repetition_bigram_ratio(out)
        unk = "<unk>" in (out or "")
        out_has_q = ("[?]" in (out or "")) or ("?" in (out or ""))
        src_has_q = "?" in (ko or "")

        src_len = len(re.sub(r"\s+", "", str(ko)))
        out_len = len(re.sub(r"\s+", "", str(out)))
        len_ratio_src = out_len / max(1, src_len)
        ref_len = len(re.sub(r"\s+", "", str(ref))) if ref else 0
        len_ratio_ref = (out_len / max(1, ref_len)) if ref_len else None

        rows.append(
            {
                "ko": ko,
                "ref": ref,
                "out": out,
                "unk": int(unk),
                "src_qmark": int(src_has_q),
                "out_qmark": int(out_has_q),
                "keep_score": keep_score,
                "rep_bigram": rep_ratio,
                "max_run": max_run,
                "src_len": src_len,
                "out_len": out_len,
                "len_ratio_src": len_ratio_src,
                "len_ratio_ref": len_ratio_ref,
            }
        )
    t1 = time.perf_counter()

    print("\n[Eval 总结]")
    print("translator:", translator_path)
    print("model:", model_path)
    print("spm_ko:", ko_spm_path)
    print("spm_zh:", zh_spm_path)
    print("samples:", len(rows))
    print("device:", str(device))
    print("elapsed_sec:", round(t1 - t0, 3))
    print("unk rate:", sum(r["unk"] for r in rows) / max(1, len(rows)))
    print("out '?' rate:", sum(r["out_qmark"] for r in rows) / max(1, len(rows)))
    print("avg keep(ascii/digit) score:", sum(r["keep_score"] for r in rows) / max(1, len(rows)))
    print("avg rep_bigram:", sum(r["rep_bigram"] for r in rows) / max(1, len(rows)))
    print("avg len_ratio(src):", sum(r["len_ratio_src"] for r in rows) / max(1, len(rows)))

    def collapse_flag(r):
        return r["out_len"] <= 1 or r["len_ratio_src"] < 0.25

    print("collapse rate:", sum(1 for r in rows if collapse_flag(r)) / max(1, len(rows)))

    rows_sorted = sorted(rows, key=lambda r: (collapse_flag(r), r["unk"], r["rep_bigram"], -r["keep_score"]), reverse=True)
    show = min(20, len(rows_sorted))
    print(f"\n[问题样本 Top {show}]")
    for i in range(show):
        r = rows_sorted[i]
        print("\n---")
        print("KO:", r["ko"])
        if r["ref"]:
            print("REF:", r["ref"])
        print("OUT:", r["out"])
        print(
            "unk:",
            r["unk"],
            "keep:",
            round(r["keep_score"], 2),
            "rep_bigram:",
            round(r["rep_bigram"], 2),
            "max_run:",
            r["max_run"],
            "len_ratio_src:",
            round(r["len_ratio_src"], 2),
        )

    if eval_out:
        eval_out = os.path.normpath(str(eval_out).strip().strip('"').strip("'"))
        out_wb = openpyxl.Workbook()
        ws = out_wb.active
        ws.title = "Eval"
        ws.append(
            [
                "ko",
                "ref",
                "out",
                "unk",
                "src_qmark",
                "out_qmark",
                "keep_score",
                "rep_bigram",
                "max_run",
                "src_len",
                "out_len",
                "len_ratio_src",
                "len_ratio_ref",
            ]
        )
        for r in rows:
            ws.append(
                [
                    r["ko"],
                    r["ref"],
                    r["out"],
                    r["unk"],
                    r["src_qmark"],
                    r["out_qmark"],
                    float(r["keep_score"]),
                    float(r["rep_bigram"]),
                    int(r["max_run"]),
                    int(r["src_len"]),
                    int(r["out_len"]),
                    float(r["len_ratio_src"]),
                    "" if r["len_ratio_ref"] is None else float(r["len_ratio_ref"]),
                ]
            )
        out_wb.save(eval_out)
        print("\n已保存:", eval_out)


def eval_set(
    root: str,
    eval_path: str,
    model_dir: str,
    translator_path: str,
    eval_out: str | None = None,
    eval_sheet: str | None = None,
    n: int | None = None,
    seed: int = 42,
    max_len: int = 50,
):
    tp = os.path.basename(str(translator_path or "")).lower()
    md = os.path.normpath(str(model_dir).strip().strip('"').strip("'"))
    is_bpe = ("bpe" in tp) or ("子词" in str(translator_path or "")) or os.path.exists(os.path.join(md, "spm_ko_v3_2_1.model"))
    if is_bpe:
        return eval_set_bpe(
            root=root,
            eval_path=eval_path,
            model_dir=model_dir,
            translator_path=translator_path,
            eval_out=eval_out,
            eval_sheet=eval_sheet,
            n=n,
            seed=seed,
            max_len=max_len,
        )
    return eval_set_word(
        root=root,
        eval_path=eval_path,
        model_dir=model_dir,
        translator_path=translator_path,
        eval_out=eval_out,
        eval_sheet=eval_sheet,
        n=n,
        seed=seed,
        max_len=max_len,
    )


def compare_rt(root: str, corpus_xlsx: str, model_dir: str, v30_path: str, v301_path: str, n: int, seed: int):
    try:
        import torch
    except Exception as e:
        print("缺少 torch，无法在本环境跑对比。错误:", e)
        print("请使用装有 torch 的解释器运行，比如：python diagnose_mt.py --compare")
        return

    corpus_path = os.path.join(root, corpus_xlsx)
    model_path, ko_vocab_path, zh_vocab_path = _infer_model_paths(model_dir)
    user_dict_path = os.path.join(root, "user_dict.md")

    mod30 = _load_module_from_path("rt_v3_0", v30_path)
    mod301 = _load_module_from_path("rt_v3_0_1", v301_path)

    with open(ko_vocab_path, "rb") as f:
        ko_vocab = pickle.load(f)
    with open(zh_vocab_path, "rb") as f:
        zh_vocab = pickle.load(f)

    device = torch.device("cpu")
    input_dim = len(ko_vocab)
    output_dim = len(zh_vocab)
    enc_emb_dim = 256
    dec_emb_dim = 256
    hid_dim = 512
    n_layers = 1

    attn = mod301.Attention(hid_dim)
    enc = mod301.Encoder(input_dim, enc_emb_dim, hid_dim, n_layers, 0)
    dec = mod301.Decoder(output_dim, dec_emb_dim, hid_dim, n_layers, 0, attn)
    model = mod301.Seq2Seq(enc, dec, device).to(device)
    model.load_state_dict(torch.load(model_path, map_location=device))
    model.eval()

    user_dict_30 = mod30.load_user_dict(user_dict_path)
    user_dict_301 = mod301.load_user_dict(user_dict_path)

    ko_sents, zh_sents = _read_parallel_corpus_xlsx(corpus_path, max_rows=None)
    pairs = [(k, z) for k, z in zip(ko_sents, zh_sents) if isinstance(k, str) and k.strip() and isinstance(z, str) and z.strip()]

    random.seed(seed)
    random.shuffle(pairs)
    pairs = pairs[: max(1, n)]

    rows = []
    t0 = time.perf_counter()
    for ko, ref in pairs:
        out30 = mod30.translate_sentence(ko, model, ko_vocab, zh_vocab, device, user_dict_30, show_tokens=False)
        out301 = mod301.translate_sentence(ko, model, ko_vocab, zh_vocab, device, user_dict_301, show_tokens=False)

        sim = difflib.SequenceMatcher(None, out30, out301).ratio()
        unk30 = "<unk>" in out30
        unk301 = "<unk>" in out301

        keep_src = _ascii_words_digits(ko)
        keep30 = _ascii_words_digits(out30)
        keep301 = _ascii_words_digits(out301)
        keep_src_set = set(keep_src)
        keep30_score = sum(1 for x in keep_src_set if x in set(keep30)) / max(1, len(keep_src_set))
        keep301_score = sum(1 for x in keep_src_set if x in set(keep301)) / max(1, len(keep_src_set))

        rows.append(
            {
                "ko": ko,
                "ref": ref,
                "out30": out30,
                "out301": out301,
                "sim": sim,
                "unk30": unk30,
                "unk301": unk301,
                "keep30": keep30_score,
                "keep301": keep301_score,
            }
        )
    t1 = time.perf_counter()

    print("\n[对比总结]")
    print("samples:", len(rows))
    print("elapsed_sec:", round(t1 - t0, 3))
    print("avg similarity(out30 vs out301):", sum(r["sim"] for r in rows) / max(1, len(rows)))
    print("unk rate v3.0:", sum(1 for r in rows if r["unk30"]) / max(1, len(rows)))
    print("unk rate v3.0.1:", sum(1 for r in rows if r["unk301"]) / max(1, len(rows)))
    print("avg ascii/digit keep v3.0:", sum(r["keep30"] for r in rows) / max(1, len(rows)))
    print("avg ascii/digit keep v3.0.1:", sum(r["keep301"] for r in rows) / max(1, len(rows)))

    rows_sorted = sorted(rows, key=lambda r: (r["sim"], r["unk301"] is True), reverse=False)
    show = min(15, len(rows_sorted))
    print(f"\n[差异最大的 {show} 条样本] (sim 越小差异越大)")
    for i in range(show):
        r = rows_sorted[i]
        print("\n---")
        print("KO:", r["ko"])
        print("REF:", r["ref"])
        print("V3.0 :", r["out30"])
        print("V3.0.1:", r["out301"])
        print("sim:", round(r["sim"], 3), "unk30:", r["unk30"], "unk301:", r["unk301"], "keep30:", round(r["keep30"], 2), "keep301:", round(r["keep301"], 2))


KO_RE = re.compile(r"[\uAC00-\uD7A3]")
ZH_RE = re.compile(r"[\u4E00-\u9FFF]")


def _count_regex(pattern: re.Pattern, text: str) -> int:
    if not isinstance(text, str) or not text:
        return 0
    return len(pattern.findall(text))


def clean_corpus_xlsx(
    input_xlsx: str,
    output_xlsx: str,
    sheet_name: str = "Corpus (2)",
    min_hangul_ratio: float = 0.2,
    min_hanzi_ratio: float = 0.2,
    drop_identical: bool = True,
    drop_qmark: bool = True,
    prefer_zh_col: int = 3,
):
    wb = openpyxl.load_workbook(input_xlsx, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise RuntimeError(f"找不到工作表: {sheet_name}，可用: {wb.sheetnames}")

    ws = wb[sheet_name]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    header = list(header) if header else ["序号", "KO", "ZH", "ZH修正"]

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = sheet_name
    out_ws.append(header)

    removed_ws = out_wb.create_sheet("Removed")
    removed_ws.append(["reason", *header])

    total = kept = 0
    dropped_by_reason = Counter()

    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        row_list = list(row) if row else []

        ko = row_list[1] if len(row_list) > 1 else None
        zh_pref = row_list[prefer_zh_col] if len(row_list) > prefer_zh_col else None
        zh_fallback = row_list[2] if len(row_list) > 2 else None

        ko_s = "" if ko is None else str(ko).strip()
        zh_s = "" if zh_pref is None else str(zh_pref).strip()
        if not zh_s:
            zh_s = "" if zh_fallback is None else str(zh_fallback).strip()

        if not ko_s or not zh_s:
            dropped_by_reason["empty_ko_or_zh"] += 1
            removed_ws.append(["empty_ko_or_zh", *row_list])
            continue

        if drop_qmark and ("?" in ko_s or "?" in zh_s):
            dropped_by_reason["contains_qmark"] += 1
            removed_ws.append(["contains_qmark", *row_list])
            continue

        if drop_identical and ko_s == zh_s:
            dropped_by_reason["ko_eq_zh"] += 1
            removed_ws.append(["ko_eq_zh", *row_list])
            continue

        ko_h = _count_regex(KO_RE, ko_s)
        zh_h = _count_regex(ZH_RE, zh_s)

        if ko_h == 0:
            dropped_by_reason["no_hangul"] += 1
            removed_ws.append(["no_hangul", *row_list])
            continue

        if zh_h == 0:
            dropped_by_reason["no_hanzi"] += 1
            removed_ws.append(["no_hanzi", *row_list])
            continue

        ko_ratio = ko_h / max(1, len(ko_s))
        zh_ratio = zh_h / max(1, len(zh_s))
        if ko_ratio < float(min_hangul_ratio):
            dropped_by_reason["low_hangul_ratio"] += 1
            removed_ws.append(["low_hangul_ratio", *row_list])
            continue
        if zh_ratio < float(min_hanzi_ratio):
            dropped_by_reason["low_hanzi_ratio"] += 1
            removed_ws.append(["low_hanzi_ratio", *row_list])
            continue

        kept += 1
        out_row = row_list[:]
        if len(out_row) > 1:
            out_row[1] = ko_s
        if len(out_row) > prefer_zh_col:
            out_row[prefer_zh_col] = zh_s
        out_ws.append(out_row)

    wb.close()
    out_wb.save(output_xlsx)

    print("\n[clean_corpus_xlsx]")
    print("input:", input_xlsx)
    print("sheet:", sheet_name)
    print("output:", output_xlsx)
    print("rows_total:", total)
    print("rows_kept:", kept)
    print("rows_removed:", total - kept)
    if dropped_by_reason:
        print("\n[removed reasons]")
        for k, v in dropped_by_reason.most_common():
            print(f"{k}: {v}")


class _QueueWriter:
    def __init__(self, q: "queue.Queue[str]"):
        self.q = q

    def write(self, s: str):
        if not s:
            return 0
        try:
            self.q.put(str(s))
        except Exception:
            pass
        return len(s)

    def flush(self):
        return


def ui_main():
    try:
        import tkinter as tk
        from tkinter import ttk, filedialog, messagebox
    except Exception as e:
        print("tkinter 不可用，无法启动 UI。错误:", e)
        return

    root = tk.Tk()
    root.title("diagnose_mt 工具")
    root.geometry("980x720")

    q: "queue.Queue[str]" = queue.Queue()
    running = {"value": False}

    def poll_log():
        try:
            while True:
                s = q.get_nowait()
                txt.insert("end", s)
                txt.see("end")
        except queue.Empty:
            pass
        root.after(100, poll_log)

    frame = ttk.Frame(root, padding=12)
    frame.pack(fill="both", expand=True)

    def clear_log():
        txt.delete("1.0", "end")

    def _norm_path(v: str):
        return os.path.normpath(str(v).strip().strip('"').strip("'"))

    def _browse_open_xlsx(title: str):
        return filedialog.askopenfilename(title=title, filetypes=[("Excel", "*.xlsx *.xlsm"), ("All", "*.*")])

    def _browse_save_xlsx(title: str, initialdir: str | None = None):
        return filedialog.asksaveasfilename(
            title=title,
            defaultextension=".xlsx",
            initialdir=initialdir,
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("All", "*.*")],
        )

    def _browse_open_txt(title: str):
        return filedialog.askopenfilename(title=title, filetypes=[("Text", "*.txt"), ("All", "*.*")])

    def _browse_open_py(title: str):
        return filedialog.askopenfilename(title=title, filetypes=[("Python", "*.py"), ("All", "*.*")])

    def _browse_open_any(title: str):
        return filedialog.askopenfilename(title=title, filetypes=[("All", "*.*")])

    def _browse_dir(title: str):
        return filedialog.askdirectory(title=title)

    def run_in_thread(fn, args_dict: dict):
        if running["value"]:
            messagebox.showinfo("提示", "正在运行，请等待结束")
            return
        running["value"] = True
        txt.delete("1.0", "end")

        def worker():
            old_out, old_err = sys.stdout, sys.stderr
            sys.stdout, sys.stderr = _QueueWriter(q), _QueueWriter(q)
            try:
                fn(**args_dict)
            except Exception:
                print("\n[UI] 异常：\n" + traceback.format_exc())
            finally:
                sys.stdout, sys.stderr = old_out, old_err
                running["value"] = False

        threading.Thread(target=worker, daemon=True).start()

    nb = ttk.Notebook(frame)
    nb.pack(fill="both", expand=True)

    tab1 = ttk.Frame(nb, padding=10)
    tab2 = ttk.Frame(nb, padding=10)
    tab3 = ttk.Frame(nb, padding=10)
    tab4 = ttk.Frame(nb, padding=10)
    tab5 = ttk.Frame(nb, padding=10)
    nb.add(tab1, text="打标/打分")
    nb.add(tab2, text="清洗")
    nb.add(tab3, text="评估")
    nb.add(tab4, text="对比")
    nb.add(tab5, text="词表诊断")

    v1 = {
        "input_xlsx": tk.StringVar(value=r"D:\PythonProject\NEW-CORPUS-20260602.cleaned.xlsx"),
        "output_xlsx": tk.StringVar(value=r"D:\PythonProject\NEW-CORPUS-20260602.cleaned.xlsx"),
        "same_output": tk.BooleanVar(value=True),
        "sheet": tk.StringVar(value="Corpus"),
        "keywords_file": tk.StringVar(value=r"D:\PythonProject\corpus_sources\domain_keywords_ko.txt"),
        "max_ko_len": tk.StringVar(value="128"),
        "max_zh_len": tk.StringVar(value="96"),
        "do_tag": tk.BooleanVar(value=True),
        "do_score": tk.BooleanVar(value=True),
    }

    def v1_sync_output():
        if v1["same_output"].get():
            v1["output_xlsx"].set(v1["input_xlsx"].get())

    def v1_browse_input():
        p = _browse_open_xlsx("选择输入 xlsx")
        if p:
            v1["input_xlsx"].set(p)
            v1_sync_output()

    def v1_browse_output():
        p = _browse_save_xlsx("选择输出 xlsx")
        if p:
            v1["output_xlsx"].set(p)
            v1["same_output"].set(False)

    def v1_browse_keywords():
        p = _browse_open_txt("选择关键词文件")
        if p:
            v1["keywords_file"].set(p)

    def v1_run():
        input_xlsx = _norm_path(v1["input_xlsx"].get())
        output_xlsx = _norm_path(v1["output_xlsx"].get())
        sheet = v1["sheet"].get().strip() or None
        if not input_xlsx or not os.path.exists(input_xlsx):
            messagebox.showerror("参数错误", "输入文件不存在")
            return
        if not output_xlsx:
            messagebox.showerror("参数错误", "输出文件不能为空")
            return
        if not (v1["do_tag"].get() or v1["do_score"].get()):
            messagebox.showerror("参数错误", "至少选择一个功能（打标签/打分）")
            return
        kw = _norm_path(v1["keywords_file"].get())
        try:
            mk = int(v1["max_ko_len"].get().strip())
            mz = int(v1["max_zh_len"].get().strip())
        except Exception:
            messagebox.showerror("参数错误", "max_ko_len / max_zh_len 必须是整数")
            return

        def task():
            if v1["do_tag"].get():
                print("\n[UI] 开始：打标签")
                tag_corpus_xlsx(input_xlsx=input_xlsx, output_xlsx=output_xlsx, sheet_name=sheet, keywords_path=kw)
            if v1["do_score"].get():
                print("\n[UI] 开始：打分")
                score_corpus_xlsx(
                    input_xlsx=input_xlsx, output_xlsx=output_xlsx, sheet_name=sheet, max_ko_len=mk, max_zh_len=mz
                )
            print("\n[UI] 完成")

        run_in_thread(lambda: task(), {})

    r = 0
    ttk.Label(tab1, text="输入文件").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab1, textvariable=v1["input_xlsx"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab1, text="选择", command=v1_browse_input).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab1, text="输出文件").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab1, textvariable=v1["output_xlsx"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab1, text="选择", command=v1_browse_output).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Checkbutton(tab1, text="输出=输入", variable=v1["same_output"], command=v1_sync_output).grid(
        row=r, column=1, sticky="w", padx=6
    )
    r += 1
    ttk.Label(tab1, text="Sheet").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab1, textvariable=v1["sheet"], width=24).grid(row=r, column=1, sticky="w", padx=6)
    r += 1
    ttk.Separator(tab1, orient="horizontal").grid(row=r, column=0, columnspan=3, sticky="we", pady=8)
    r += 1
    ttk.Checkbutton(tab1, text="打标签", variable=v1["do_tag"]).grid(row=r, column=0, sticky="w")
    ttk.Label(tab1, text="关键词文件").grid(row=r, column=1, sticky="w", padx=6)
    ttk.Entry(tab1, textvariable=v1["keywords_file"], width=70).grid(row=r, column=1, sticky="e", padx=6)
    ttk.Button(tab1, text="选择", command=v1_browse_keywords).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Checkbutton(tab1, text="打分", variable=v1["do_score"]).grid(row=r, column=0, sticky="w")
    sub = ttk.Frame(tab1)
    sub.grid(row=r, column=1, sticky="w", padx=6)
    ttk.Label(sub, text="max_ko_len").grid(row=0, column=0, sticky="w")
    ttk.Entry(sub, textvariable=v1["max_ko_len"], width=8).grid(row=0, column=1, sticky="w", padx=6)
    ttk.Label(sub, text="max_zh_len").grid(row=0, column=2, sticky="w")
    ttk.Entry(sub, textvariable=v1["max_zh_len"], width=8).grid(row=0, column=3, sticky="w", padx=6)
    r += 1
    ttk.Button(tab1, text="运行本页", command=v1_run).grid(row=r, column=0, sticky="w", pady=8)
    tab1.columnconfigure(1, weight=1)

    v2 = {
        "input_xlsx": tk.StringVar(value=r"D:\PythonProject\NEW-CORPUS-20260602.cleaned.xlsx"),
        "output_xlsx": tk.StringVar(value=r"D:\PythonProject\NEW-CORPUS-20260602.cleaned.xlsx"),
        "same_output": tk.BooleanVar(value=False),
        "sheet": tk.StringVar(value="Corpus"),
        "min_hangul_ratio": tk.StringVar(value="0.2"),
        "min_hanzi_ratio": tk.StringVar(value="0.2"),
        "keep_identical": tk.BooleanVar(value=False),
        "keep_qmark": tk.BooleanVar(value=False),
        "prefer_zh_col": tk.StringVar(value="3"),
    }

    def v2_sync_output():
        if v2["same_output"].get():
            v2["output_xlsx"].set(v2["input_xlsx"].get())

    def v2_browse_input():
        p = _browse_open_xlsx("选择输入 xlsx")
        if p:
            v2["input_xlsx"].set(p)
            v2_sync_output()

    def v2_browse_output():
        p = _browse_save_xlsx("选择输出 xlsx")
        if p:
            v2["output_xlsx"].set(p)
            v2["same_output"].set(False)

    def v2_run():
        input_xlsx = _norm_path(v2["input_xlsx"].get())
        output_xlsx = _norm_path(v2["output_xlsx"].get())
        sheet = v2["sheet"].get().strip() or "Corpus (2)"
        if not input_xlsx or not os.path.exists(input_xlsx):
            messagebox.showerror("参数错误", "输入文件不存在")
            return
        if not output_xlsx:
            messagebox.showerror("参数错误", "输出文件不能为空")
            return
        try:
            mh = float(v2["min_hangul_ratio"].get().strip())
            mz = float(v2["min_hanzi_ratio"].get().strip())
            pref = int(v2["prefer_zh_col"].get().strip())
        except Exception:
            messagebox.showerror("参数错误", "比例必须是数字，prefer_zh_col 必须是整数")
            return

        def task():
            print("\n[UI] 开始：清洗")
            clean_corpus_xlsx(
                input_xlsx=input_xlsx,
                output_xlsx=output_xlsx,
                sheet_name=sheet,
                min_hangul_ratio=mh,
                min_hanzi_ratio=mz,
                drop_identical=(not v2["keep_identical"].get()),
                drop_qmark=(not v2["keep_qmark"].get()),
                prefer_zh_col=pref,
            )
            print("\n[UI] 完成")

        run_in_thread(lambda: task(), {})

    r = 0
    ttk.Label(tab2, text="输入文件").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab2, textvariable=v2["input_xlsx"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab2, text="选择", command=v2_browse_input).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab2, text="输出文件").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab2, textvariable=v2["output_xlsx"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab2, text="选择", command=v2_browse_output).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Checkbutton(tab2, text="输出=输入", variable=v2["same_output"], command=v2_sync_output).grid(
        row=r, column=1, sticky="w", padx=6
    )
    r += 1
    ttk.Label(tab2, text="Sheet").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab2, textvariable=v2["sheet"], width=24).grid(row=r, column=1, sticky="w", padx=6)
    r += 1
    sub = ttk.Frame(tab2)
    sub.grid(row=r, column=0, columnspan=3, sticky="w", pady=6)
    ttk.Label(sub, text="min_hangul_ratio").grid(row=0, column=0, sticky="w")
    ttk.Entry(sub, textvariable=v2["min_hangul_ratio"], width=8).grid(row=0, column=1, sticky="w", padx=6)
    ttk.Label(sub, text="min_hanzi_ratio").grid(row=0, column=2, sticky="w")
    ttk.Entry(sub, textvariable=v2["min_hanzi_ratio"], width=8).grid(row=0, column=3, sticky="w", padx=6)
    ttk.Label(sub, text="prefer_zh_col").grid(row=0, column=4, sticky="w")
    ttk.Entry(sub, textvariable=v2["prefer_zh_col"], width=6).grid(row=0, column=5, sticky="w", padx=6)
    r += 1
    ttk.Checkbutton(tab2, text="保留 KO==ZH", variable=v2["keep_identical"]).grid(row=r, column=0, sticky="w")
    ttk.Checkbutton(tab2, text="保留 ? 行", variable=v2["keep_qmark"]).grid(row=r, column=1, sticky="w")
    r += 1
    ttk.Button(tab2, text="运行本页", command=v2_run).grid(row=r, column=0, sticky="w", pady=8)
    tab2.columnconfigure(1, weight=1)

    v3 = {
        "eval_path": tk.StringVar(value=r"D:\PythonProject\for_live_sentences_ko.txt"),
        "model_dir": tk.StringVar(value=r"D:\PythonProject\Translate Model\V3.0(Attention)\20260527-best Model\epoch03-4.4666"),
        "translator": tk.StringVar(value=r"D:\PythonProject\实时翻译测试_V3.0.1(greedy).py"),
        "eval_out": tk.StringVar(value=r"D:\PythonProject\evaluate models performance\evaluation_result.xlsx"),
        "eval_sheet": tk.StringVar(value=""),
        "n": tk.StringVar(value="100"),
        "seed": tk.StringVar(value="42"),
        "max_len": tk.StringVar(value="80"),
    }

    def v3_browse_eval():
        p = _browse_open_any("选择评估集文件")
        if p:
            v3["eval_path"].set(p)

    def v3_browse_model_dir():
        p = _browse_dir("选择模型目录")
        if p:
            v3["model_dir"].set(p)

    def v3_browse_translator():
        p = _browse_open_py("选择翻译脚本")
        if p:
            v3["translator"].set(p)

    def v3_browse_eval_out():
        # 设置默认打开的目录
        init_dir = r"D:\PythonProject\evaluate models performance"
        if not os.path.exists(init_dir):
            try:
                os.makedirs(init_dir, exist_ok=True)
            except Exception:
                init_dir = None
        
        p = _browse_save_xlsx("选择导出评估表 xlsx", initialdir=init_dir)
        if p:
            v3["eval_out"].set(p)

    def v3_run():
        eval_path = _norm_path(v3["eval_path"].get())
        model_dir = _norm_path(v3["model_dir"].get())
        translator = _norm_path(v3["translator"].get())
        eval_out = _norm_path(v3["eval_out"].get()) if v3["eval_out"].get().strip() else None
        
        # 确保评估结果目录存在
        if eval_out:
            eval_out_dir = os.path.dirname(eval_out)
            if eval_out_dir and not os.path.exists(eval_out_dir):
                try:
                    os.makedirs(eval_out_dir, exist_ok=True)
                except Exception as e:
                    messagebox.showwarning("目录创建失败", f"无法创建目录: {eval_out_dir}\n错误: {e}")
        eval_sheet = v3["eval_sheet"].get().strip() or None
        try:
            n = int(v3["n"].get().strip())
            seed = int(v3["seed"].get().strip())
            max_len = int(v3["max_len"].get().strip())
        except Exception:
            messagebox.showerror("参数错误", "n/seed/max_len 必须是整数")
            return
        if not eval_path or not os.path.exists(eval_path):
            messagebox.showerror("参数错误", "评估集文件不存在")
            return
        if not model_dir or not os.path.isdir(model_dir):
            messagebox.showerror("参数错误", "model_dir 不是目录")
            return
        if not translator or not os.path.exists(translator):
            messagebox.showerror("参数错误", "翻译脚本不存在")
            return

        def task():
            print("\n[UI] 开始：评估")
            eval_set(
                root=os.path.dirname(__file__),
                eval_path=eval_path,
                model_dir=model_dir,
                translator_path=translator,
                eval_out=eval_out,
                eval_sheet=eval_sheet,
                n=n,
                seed=seed,
                max_len=max_len,
            )
            print("\n[UI] 完成")

        run_in_thread(lambda: task(), {})

    r = 0
    ttk.Label(tab3, text="评估集文件").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab3, textvariable=v3["eval_path"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab3, text="选择", command=v3_browse_eval).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab3, text="model_dir").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab3, textvariable=v3["model_dir"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab3, text="选择", command=v3_browse_model_dir).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab3, text="translator.py").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab3, textvariable=v3["translator"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab3, text="选择", command=v3_browse_translator).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab3, text="eval_out(可选)").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab3, textvariable=v3["eval_out"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab3, text="选择", command=v3_browse_eval_out).grid(row=r, column=2, sticky="we")
    r += 1
    sub = ttk.Frame(tab3)
    sub.grid(row=r, column=0, columnspan=3, sticky="w", pady=6)
    ttk.Label(sub, text="sheet(可选)").grid(row=0, column=0, sticky="w")
    ttk.Entry(sub, textvariable=v3["eval_sheet"], width=12).grid(row=0, column=1, sticky="w", padx=6)
    ttk.Label(sub, text="n").grid(row=0, column=2, sticky="w")
    ttk.Entry(sub, textvariable=v3["n"], width=8).grid(row=0, column=3, sticky="w", padx=6)
    ttk.Label(sub, text="seed").grid(row=0, column=4, sticky="w")
    ttk.Entry(sub, textvariable=v3["seed"], width=8).grid(row=0, column=5, sticky="w", padx=6)
    ttk.Label(sub, text="max_len").grid(row=0, column=6, sticky="w")
    ttk.Entry(sub, textvariable=v3["max_len"], width=8).grid(row=0, column=7, sticky="w", padx=6)
    r += 1
    ttk.Button(tab3, text="运行本页", command=v3_run).grid(row=r, column=0, sticky="w", pady=8)
    tab3.columnconfigure(1, weight=1)

    v4 = {
        "corpus_xlsx": tk.StringVar(value=r"D:\PythonProject\Corpus(K2C)-2.xlsx"),
        "model_dir": tk.StringVar(value=r"D:\PythonProject\Translate Model\V3.0(Attention)\20260527-best Model\epoch03-4.4666"),
        "v30": tk.StringVar(value=r"D:\PythonProject\实时翻译测试_V3.0(greedy).py"),
        "v301": tk.StringVar(value=r"D:\PythonProject\实时翻译测试_V3.0.1(greedy).py"),
        "n": tk.StringVar(value="30"),
        "seed": tk.StringVar(value="42"),
    }

    def v4_browse_corpus():
        p = _browse_open_xlsx("选择语料 xlsx")
        if p:
            v4["corpus_xlsx"].set(p)

    def v4_browse_model_dir():
        p = _browse_dir("选择模型目录")
        if p:
            v4["model_dir"].set(p)

    def v4_browse_v30():
        p = _browse_open_py("选择 V3.0 翻译脚本")
        if p:
            v4["v30"].set(p)

    def v4_browse_v301():
        p = _browse_open_py("选择 V3.0.1 翻译脚本")
        if p:
            v4["v301"].set(p)

    def v4_run():
        corpus_path = _norm_path(v4["corpus_xlsx"].get())
        model_dir = _norm_path(v4["model_dir"].get())
        v30 = _norm_path(v4["v30"].get())
        v301 = _norm_path(v4["v301"].get())
        try:
            n = int(v4["n"].get().strip())
            seed = int(v4["seed"].get().strip())
        except Exception:
            messagebox.showerror("参数错误", "n/seed 必须是整数")
            return
        if not corpus_path or not os.path.exists(corpus_path):
            messagebox.showerror("参数错误", "语料文件不存在")
            return
        if not model_dir or not os.path.isdir(model_dir):
            messagebox.showerror("参数错误", "model_dir 不是目录")
            return
        if not v30 or not os.path.exists(v30) or not v301 or not os.path.exists(v301):
            messagebox.showerror("参数错误", "翻译脚本不存在")
            return

        root_dir = os.path.dirname(corpus_path)
        corpus_name = os.path.basename(corpus_path)

        def task():
            print("\n[UI] 开始：对比")
            compare_rt(
                root=root_dir,
                corpus_xlsx=corpus_name,
                model_dir=model_dir,
                v30_path=v30,
                v301_path=v301,
                n=n,
                seed=seed,
            )
            print("\n[UI] 完成")

        run_in_thread(lambda: task(), {})

    r = 0
    ttk.Label(tab4, text="语料文件").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab4, textvariable=v4["corpus_xlsx"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab4, text="选择", command=v4_browse_corpus).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab4, text="model_dir").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab4, textvariable=v4["model_dir"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab4, text="选择", command=v4_browse_model_dir).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab4, text="V3.0 脚本").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab4, textvariable=v4["v30"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab4, text="选择", command=v4_browse_v30).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab4, text="V3.0.1 脚本").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab4, textvariable=v4["v301"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab4, text="选择", command=v4_browse_v301).grid(row=r, column=2, sticky="we")
    r += 1
    sub = ttk.Frame(tab4)
    sub.grid(row=r, column=0, columnspan=3, sticky="w", pady=6)
    ttk.Label(sub, text="n").grid(row=0, column=0, sticky="w")
    ttk.Entry(sub, textvariable=v4["n"], width=8).grid(row=0, column=1, sticky="w", padx=6)
    ttk.Label(sub, text="seed").grid(row=0, column=2, sticky="w")
    ttk.Entry(sub, textvariable=v4["seed"], width=8).grid(row=0, column=3, sticky="w", padx=6)
    r += 1
    ttk.Button(tab4, text="运行本页", command=v4_run).grid(row=r, column=0, sticky="w", pady=8)
    tab4.columnconfigure(1, weight=1)

    v5 = {
        "corpus_xlsx": tk.StringVar(value=r"D:\PythonProject\Corpus(K2C)-2.xlsx"),
        "model_dir": tk.StringVar(value=r"D:\PythonProject\Translate Model\V3.0(Attention)\20260527-best Model\epoch03-4.4666"),
    }

    def v5_browse_corpus():
        p = _browse_open_xlsx("选择语料 xlsx")
        if p:
            v5["corpus_xlsx"].set(p)

    def v5_browse_model_dir():
        p = _browse_dir("选择模型目录")
        if p:
            v5["model_dir"].set(p)

    def v5_run():
        corpus_path = _norm_path(v5["corpus_xlsx"].get())
        model_dir = _norm_path(v5["model_dir"].get())
        if not corpus_path or not os.path.exists(corpus_path):
            messagebox.showerror("参数错误", "语料文件不存在")
            return
        if not model_dir or not os.path.isdir(model_dir):
            messagebox.showerror("参数错误", "model_dir 不是目录")
            return
        root_dir = os.path.dirname(corpus_path)
        corpus_name = os.path.basename(corpus_path)

        def task():
            print("\n[UI] 开始：词表诊断")
            vocab_diagnose(root=root_dir, corpus_xlsx=corpus_name, model_dir=model_dir)
            print("\n[UI] 完成")

        run_in_thread(lambda: task(), {})

    r = 0
    ttk.Label(tab5, text="语料文件").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab5, textvariable=v5["corpus_xlsx"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab5, text="选择", command=v5_browse_corpus).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Label(tab5, text="model_dir").grid(row=r, column=0, sticky="w")
    ttk.Entry(tab5, textvariable=v5["model_dir"], width=86).grid(row=r, column=1, sticky="we", padx=6)
    ttk.Button(tab5, text="选择", command=v5_browse_model_dir).grid(row=r, column=2, sticky="we")
    r += 1
    ttk.Button(tab5, text="运行本页", command=v5_run).grid(row=r, column=0, sticky="w", pady=8)
    tab5.columnconfigure(1, weight=1)

    bottom = ttk.Frame(frame)
    bottom.pack(fill="both", expand=False, pady=(10, 0))
    btns = ttk.Frame(bottom)
    btns.pack(fill="x")
    ttk.Button(btns, text="清空输出", command=clear_log).pack(side="left")
    txt = tk.Text(bottom, height=18, wrap="word")
    txt.pack(fill="both", expand=True, pady=(8, 0))

    poll_log()
    root.mainloop()


def main():
    _configure_console_utf8()
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", default=r"D:\PythonProject")
    parser.add_argument("--corpus", default="Corpus(K2C)-2.xlsx")
    parser.add_argument("--model-dir", default=None)
    parser.add_argument("--compare", action="store_true")
    parser.add_argument("--clean-corpus", action="store_true")
    parser.add_argument("--tag-corpus", action="store_true")
    parser.add_argument("--score-corpus", action="store_true")
    parser.add_argument("--summarize-corpus", action="store_true")
    parser.add_argument("--keywords-file", default=None)
    parser.add_argument("--max-ko-len", type=int, default=128)
    parser.add_argument("--max-zh-len", type=int, default=96)
    parser.add_argument("--eval-set", default=None)
    parser.add_argument("--eval-out", default=None)
    parser.add_argument("--eval-sheet", default=None)
    parser.add_argument("--translator", default=None)
    parser.add_argument("--max-len", type=int, default=50)
    parser.add_argument("--input-xlsx", default=None)
    parser.add_argument("--output-xlsx", default=None)
    parser.add_argument("--sheet", default="Corpus (2)")
    parser.add_argument("--min-hangul-ratio", type=float, default=0.2)
    parser.add_argument("--min-hanzi-ratio", type=float, default=0.2)
    parser.add_argument("--keep-identical", action="store_true")
    parser.add_argument("--keep-qmark", action="store_true")
    parser.add_argument("--n", type=int, default=30)
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--v30", default=None)
    parser.add_argument("--v301", default=None)
    parser.add_argument("--ui", action="store_true")
    args = parser.parse_args()

    model_dir = args.model_dir or os.path.join(args.root, "Translate Model")

    if args.ui or len(sys.argv) == 1:
        ui_main()
        return

    if args.summarize_corpus:
        input_xlsx = args.input_xlsx or os.path.join(args.root, "NEW-CORPUS-20260604.xlsx")
        summarize_corpus_xlsx(input_xlsx=input_xlsx, sheet_name=args.sheet if args.sheet else None)
        return

    if args.score_corpus:
        input_xlsx = args.input_xlsx or os.path.join(args.root, "Corpus(K2C)-2_20260526.xlsx")
        output_xlsx = args.output_xlsx or input_xlsx
        score_corpus_xlsx(
            input_xlsx=input_xlsx,
            output_xlsx=output_xlsx,
            sheet_name=args.sheet if args.sheet else None,
            max_ko_len=int(args.max_ko_len),
            max_zh_len=int(args.max_zh_len),
        )
        return

    if args.tag_corpus:
        input_xlsx = args.input_xlsx or os.path.join(args.root, "Corpus(K2C)-2_20260526.xlsx")
        if args.output_xlsx:
            output_xlsx = args.output_xlsx
        else:
            base, ext = os.path.splitext(input_xlsx)
            output_xlsx = base + ".tagged.xlsx"
        keywords_file = args.keywords_file or os.path.join(args.root, "corpus_sources", "domain_keywords_ko.txt")
        tag_corpus_xlsx(
            input_xlsx=input_xlsx,
            output_xlsx=output_xlsx,
            sheet_name=args.sheet if args.sheet else None,
            keywords_path=keywords_file,
        )
        return

    if args.eval_set:
        translator_path = args.translator or os.path.join(args.root, "实时翻译测试_V3.0.1(greedy).py")
        eval_set(
            root=args.root,
            eval_path=args.eval_set,
            model_dir=model_dir,
            translator_path=translator_path,
            eval_out=args.eval_out,
            eval_sheet=args.eval_sheet,
            n=args.n,
            seed=args.seed,
            max_len=int(args.max_len),
        )
        return

    if args.clean_corpus:
        input_xlsx = args.input_xlsx or os.path.join(args.root, "NEW-CORPUS-20260525.xlsx")
        output_xlsx = args.output_xlsx or os.path.join(args.root, "NEW-CORPUS-20260525.cleaned.xlsx")
        clean_corpus_xlsx(
            input_xlsx=input_xlsx,
            output_xlsx=output_xlsx,
            sheet_name=args.sheet,
            min_hangul_ratio=float(args.min_hangul_ratio),
            min_hanzi_ratio=float(args.min_hanzi_ratio),
            drop_identical=(not args.keep_identical),
            drop_qmark=(not args.keep_qmark),
        )
        return

    if args.compare:
        v30_path = args.v30 or os.path.join(args.root, "实时翻译测试_V3.0(greedy).py")
        v301_path = args.v301 or os.path.join(args.root, "实时翻译测试_V3.0.1(greedy).py")
        compare_rt(
            root=args.root,
            corpus_xlsx=args.corpus,
            model_dir=model_dir,
            v30_path=v30_path,
            v301_path=v301_path,
            n=args.n,
            seed=args.seed,
        )
        return

    vocab_diagnose(root=args.root, corpus_xlsx=args.corpus, model_dir=model_dir)


if __name__ == "__main__":
    main()
